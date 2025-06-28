from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import os
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)  # Enable CORS for frontend requests

# Configuration
UPLOAD_FOLDER = 'uploads'
EXCEL_FOLDER = 'excel_files'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)

# Email configuration (replace with your email settings)
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'email': 'btdsahil7@gmail.com',  # Replace with your email
    'password': 'lvjlzyanbaegslto',   # Replace with your app password
    'recipient': 'sahilsingh0532@gmail.com'
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def create_styled_excel(df, filename):
    """Create a well-formatted Excel file with styling"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    return filename

def send_email_with_attachment(subject, body, attachment_path, recipient_email):
    """Send email with Excel attachment"""
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['email']
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # Add body to email
        msg.attach(MIMEText(body, 'html'))
        
        # Add attachment
        if os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        
        # Connect to server and send email
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['email'], EMAIL_CONFIG['password'])
        text = msg.as_string()
        server.sendmail(EMAIL_CONFIG['email'], recipient_email, text)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Email sending failed: {str(e)}")
        return False

@app.route('/api/submit-id-application', methods=['POST'])
def submit_application():
    try:
        # Get form data
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['fullName', 'rollNumber', 'year', 'branch', 'email', 'phone']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create application record
        application_data = {
            'Application ID': data.get('applicationId', f"APP{datetime.now().strftime('%Y%m%d%H%M%S')}"),
            'Full Name': data['fullName'],
            'Roll Number': data['rollNumber'],
            'Year': data['year'],
            'Branch': data['branch'],
            'Email': data['email'],
            'Phone': data['phone'],
            'Submission Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Status': 'Pending'
        }
        
        # Create filename based on year and branch
        sheet_name = f"{data['branch']}_{data['year']}_Year"
        excel_filename = os.path.join(EXCEL_FOLDER, f"{sheet_name}.xlsx")
        
        # Check if Excel file exists, if not create new DataFrame
        if os.path.exists(excel_filename):
            try:
                existing_df = pd.read_excel(excel_filename)
                # Append new data
                new_df = pd.concat([existing_df, pd.DataFrame([application_data])], ignore_index=True)
            except Exception as e:
                # If file is corrupted or can't be read, create new DataFrame
                new_df = pd.DataFrame([application_data])
        else:
            new_df = pd.DataFrame([application_data])
        
        # Create styled Excel file
        create_styled_excel(new_df, excel_filename)
        
        # Prepare email content
        email_subject = "New ID Card Registration Submission"
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
                    🆔 New ID Card Application Received
                </h2>
                
                <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin: 20px 0;">
                    <h3 style="color: #2c3e50; margin-top: 0;">Student Details:</h3>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; font-weight: bold; color: #666;">Name:</td>
                            <td style="padding: 8px;">{data['fullName']}</td>
                        </tr>
                        <tr style="background: #fff;">
                            <td style="padding: 8px; font-weight: bold; color: #666;">Roll Number:</td>
                            <td style="padding: 8px;">{data['rollNumber']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold; color: #666;">Year:</td>
                            <td style="padding: 8px;">{data['year']}</td>
                        </tr>
                        <tr style="background: #fff;">
                            <td style="padding: 8px; font-weight: bold; color: #666;">Branch:</td>
                            <td style="padding: 8px;">{data['branch']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold; color: #666;">Email:</td>
                            <td style="padding: 8px;">{data['email']}</td>
                        </tr>
                        <tr style="background: #fff;">
                            <td style="padding: 8px; font-weight: bold; color: #666;">Phone:</td>
                            <td style="padding: 8px;">{data['phone']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold; color: #666;">Submission Time:</td>
                            <td style="padding: 8px;">{application_data['Submission Date']}</td>
                        </tr>
                    </table>
                </div>
                
                <div style="background: #e8f4f8; padding: 15px; border-radius: 10px; border-left: 4px solid #3498db;">
                    <p style="margin: 0;"><strong>📊 Excel file attached:</strong> {sheet_name}.xlsx</p>
                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #666;">
                        The application has been automatically added to the respective year and branch sheet.
                    </p>
                </div>
                
                <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd;">
                    <p style="color: #666; font-size: 14px;">
                        This email was automatically generated by the Tech University Xerox Center system.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Send email with attachment
        email_sent = send_email_with_attachment(
            email_subject,
            email_body,
            excel_filename,
            EMAIL_CONFIG['recipient']
        )
        
        # Prepare response
        response_data = {
            'success': True,
            'message': 'Application submitted successfully',
            'applicationId': application_data['Application ID'],
            'emailSent': email_sent
        }
        
        # Also send confirmation email to student
        if email_sent:
            student_subject = "ID Card Application Confirmation"
            student_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #27ae60;">✅ Application Submitted Successfully!</h2>
                    
                    <p>Dear {data['fullName']},</p>
                    
                    <p>Your ID card application has been received and is being processed.</p>
                    
                    <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin: 20px 0;">
                        <p><strong>Application ID:</strong> {application_data['Application ID']}</p>
                        <p><strong>Processing Time:</strong> 24-48 hours</p>
                        <p><strong>Collection Fee:</strong> ₹50 (to be paid during collection)</p>
                    </div>
                    
                    <p><strong>Next Steps:</strong></p>
                    <ol>
                        <li>Wait for processing (24-48 hours)</li>
                        <li>Visit the Xerox Center with this confirmation</li>
                        <li>Pay ₹50 and collect your ID card</li>
                    </ol>
                    
                    <p>Thank you for using our services!</p>
                    
                    <p style="color: #666; font-size: 14px; margin-top: 30px;">
                        Tech University Xerox Center<br>
                        Ground Floor, Student Center
                    </p>
                </div>
            </body>
            </html>
            """
            
            send_email_with_attachment(
                student_subject,
                student_body,
                None,  # No attachment for student
                data['email']
            )
        
        return jsonify(response_data), 200
        
    except Exception as e:
        print(f"Error processing application: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error occurred while processing your application'
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'message': 'ID Card Application API is running'
    })

@app.route('/api/applications/<branch>/<year>', methods=['GET'])
def get_applications(branch, year):
    """Get applications for specific branch and year"""
    try:
        sheet_name = f"{branch}_{year}_Year"
        excel_filename = os.path.join(EXCEL_FOLDER, f"{sheet_name}.xlsx")
        
        if os.path.exists(excel_filename):
            df = pd.read_excel(excel_filename)
            return jsonify({
                'success': True,
                'data': df.to_dict('records'),
                'count': len(df)
            })
        else:
            return jsonify({
                'success': True,
                'data': [],
                'count': 0,
                'message': 'No applications found for this branch and year'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    print("🚀 Starting ID Card Application Backend Server...")
    print("📧 Make sure to update EMAIL_CONFIG with your email credentials")
    print("🔗 Frontend should make requests to: http://localhost:5000/api/")
    print("\nAvailable endpoints:")
    print("  POST /api/submit-id-application - Submit new application")
    print("  GET  /api/health - Health check")
    print("  GET  /api/applications/<branch>/<year> - Get applications by branch and year")
    
    app.run(debug=True, host='0.0.0.0', port=5000)